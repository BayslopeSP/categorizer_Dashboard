import pandas as pd
import json
import os
import numpy as np
import openpyxl

def convert():
    files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    if not files: return print("Error: No .xlsx file found")
    
    file_path = files[0]
    print(f"Reading {file_path} with links...")

    # Load workbook to extract hyperlinks
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Step 1: Find header row
    header_row_index = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if "Sr. No." in [str(v).strip() for v in row if v]:
            header_row_index = i
            break

    # Step 2: Extract Headers
    headers = [str(cell.value).strip().replace('\n', ' ') for cell in ws[header_row_index]]
    pub_col_idx = next((i for i, h in enumerate(headers) if "Publication" in h), None)

    # Step 3: Category Mapping (Row above headers)
    cat_row = ws[header_row_index - 1]
    category_mapping = {}
    current_cat = None
    for i, cell in enumerate(cat_row):
        val = str(cell.value).strip() if cell.value else None
        if val and val != 'None': current_cat = val
        sub_val = headers[i]
        if current_cat and sub_val: category_mapping[sub_val] = current_cat

    # Step 4: Extract Data and Links
    data = []
    for row in ws.iter_rows(min_row=header_row_index + 1):
        row_dict = {}
        for i, cell in enumerate(row):
            col_name = headers[i]
            val = cell.value
            
            # If this is the Publication Number column, check for hyperlink
            if i == pub_col_idx and cell.hyperlink:
                row_dict[f"{col_name}_link"] = cell.hyperlink.target
            
            row_dict[col_name] = val
        
        # Only add if Publication Number is not empty
        if pub_col_idx is not None and row[pub_col_idx].value:
            data.append(row_dict)

    # Convert to valid JSON format
    output = {
        "mapping": category_mapping,
        "data": data
    }
    
    with open("public/data.json", "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=lambda x: None)

    print(f"Done ✅ {len(data)} records with links saved!")

if __name__ == "__main__":
    convert()